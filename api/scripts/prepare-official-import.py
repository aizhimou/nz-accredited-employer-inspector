#!/usr/bin/env python3
"""Validate an MBIE OIA workbook and generate an idempotent D1 SQL import."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date, datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import sys
import time
import unicodedata

try:
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - depends on the caller's environment
    raise SystemExit(
        "openpyxl is required. Run: python3 -m pip install -r requirements-import.txt"
    ) from error


EXPECTED_HEADERS = (
    "Organisation Name",
    "Trading Name",
    "NZBN",
    "Accreditation Type",
    "Accreditation Status",
    "Sector",
    "Subsector",
    "Expiry Date",
    "Start Date",
    "Region",
    "City",
)
SNAPSHOT_PATTERN = re.compile(
    r"List of Accredited Employers as at (\d{1,2} [A-Za-z]+ \d{4})",
    re.IGNORECASE,
)
NZBN_PATTERN = re.compile(r"\d{13}")
CONTROL_CHARACTER_PATTERN = re.compile(r"[\x00-\x1f\x7f-\x9f]")
TRADING_NAME_NULLS = {"", "-", "n/a", "na", "none", "null"}
NULL_TEXT = {"", "null"}
MAX_SQL_STATEMENT_BYTES = 100_000
DEFAULT_BATCH_SIZE = 100


@dataclass(frozen=True)
class ImportRow:
    source_row_number: int
    employer_name: str
    normalized_employer_name: str
    trading_name: str | None
    normalized_trading_name: str | None
    nzbn: str | None
    accreditation_type: str
    accreditation_status: str
    sector: str | None
    subsector: str | None
    expiry_date: str
    start_date: str
    region: str | None
    city: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the MBIE .xlsx appendix")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(".generated/official-employer-import.sql"),
        help="Generated SQL path (default: .generated/official-employer-import.sql)",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=DEFAULT_BATCH_SIZE,
        help=f"Rows per INSERT statement (default: {DEFAULT_BATCH_SIZE})",
    )
    return parser.parse_args()


def text(value: object, field_name: str, *, nullable: bool = False) -> str | None:
    if value is None:
        if nullable:
            return None
        raise ValueError(f"{field_name} is missing")
    result = str(value).strip()
    if result == "" or CONTROL_CHARACTER_PATTERN.search(result):
        if nullable and result == "":
            return None
        raise ValueError(f"{field_name} is invalid")
    return result


def nullable_text(value: object, null_values: set[str] = NULL_TEXT) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    if result.lower() in null_values:
        return None
    if CONTROL_CHARACTER_PATTERN.search(result):
        raise ValueError("text contains a control character")
    return result


def normalize_name(value: str) -> str:
    return " ".join(unicodedata.normalize("NFKC", value).strip().split()).lower()


def parse_nzbn(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError("NZBN cannot be a boolean")
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f"NZBN is not an integer: {value}")
        result = str(int(value))
    elif isinstance(value, int):
        result = str(value)
    else:
        result = str(value).strip()
    if result.lower() == "null":
        return None
    if NZBN_PATTERN.fullmatch(result) is None:
        raise ValueError(f"NZBN must contain exactly 13 digits: {result}")
    return result


def parse_local_datetime(value: object, field_name: str) -> str:
    parsed: date
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = date.fromisoformat(value.strip()[:10])
        except ValueError as error:
            raise ValueError(f"{field_name} is not an ISO date: {value}") from error
    else:
        raise ValueError(f"{field_name} is not a date")
    return f"{parsed.isoformat()}T00:00:00"


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def statement(sql: str) -> str:
    result = sql.strip() + ";\n"
    size = len(result.encode("utf-8"))
    if size > MAX_SQL_STATEMENT_BYTES:
        raise ValueError(
            f"Generated SQL statement is {size:,} bytes; D1 permits {MAX_SQL_STATEMENT_BYTES:,}"
        )
    return result


def read_workbook(path: Path) -> tuple[date, list[ImportRow], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != ["Data"]:
        raise ValueError(f"Expected one sheet named Data, found: {workbook.sheetnames}")
    sheet = workbook["Data"]

    snapshot_date: date | None = None
    header_row_number: int | None = None
    for row_number, values in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = tuple("" if value is None else str(value).strip() for value in values)
        for cell in cells:
            match = SNAPSHOT_PATTERN.search(cell)
            if match is not None:
                snapshot_date = datetime.strptime(match.group(1), "%d %B %Y").date()
        if cells[: len(EXPECTED_HEADERS)] == EXPECTED_HEADERS:
            header_row_number = row_number
            break

    if snapshot_date is None:
        raise ValueError("Could not find the official snapshot date in the workbook preamble")
    if header_row_number is None:
        raise ValueError("Could not find the expected 11-column header row")

    rows: list[ImportRow] = []
    seen_source_rows: set[int] = set()
    seen_nzbns: dict[str, int] = {}
    for source_row_number, values in enumerate(
        sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
        start=header_row_number + 1,
    ):
        source_values = values[: len(EXPECTED_HEADERS)]
        if all(value is None or str(value).strip() == "" for value in source_values):
            continue
        try:
            employer_name = text(source_values[0], "Organisation Name")
            assert employer_name is not None
            trading_name = nullable_text(source_values[1], TRADING_NAME_NULLS)
            nzbn = parse_nzbn(source_values[2])
            accreditation_type = text(source_values[3], "Accreditation Type")
            accreditation_status = text(source_values[4], "Accreditation Status")
            assert accreditation_type is not None and accreditation_status is not None
            row = ImportRow(
                source_row_number=source_row_number,
                employer_name=employer_name,
                normalized_employer_name=normalize_name(employer_name),
                trading_name=trading_name,
                normalized_trading_name=(
                    None if trading_name is None else normalize_name(trading_name)
                ),
                nzbn=nzbn,
                accreditation_type=accreditation_type,
                accreditation_status=accreditation_status,
                sector=nullable_text(source_values[5]),
                subsector=nullable_text(source_values[6]),
                expiry_date=parse_local_datetime(source_values[7], "Expiry Date"),
                start_date=parse_local_datetime(source_values[8], "Start Date"),
                region=nullable_text(source_values[9]),
                city=nullable_text(source_values[10]),
            )
        except ValueError as error:
            raise ValueError(f"Workbook row {source_row_number}: {error}") from error

        if source_row_number in seen_source_rows:
            raise ValueError(f"Duplicate source row number: {source_row_number}")
        seen_source_rows.add(source_row_number)
        if nzbn is not None:
            prior_row = seen_nzbns.get(nzbn)
            if prior_row is not None:
                raise ValueError(
                    f"NZBN {nzbn} is duplicated on workbook rows {prior_row} and {source_row_number}"
                )
            seen_nzbns[nzbn] = source_row_number
        rows.append(row)

    if not rows:
        raise ValueError("The workbook contains no employer rows")
    return snapshot_date, rows, header_row_number


def row_values(snapshot_date: str, row: ImportRow) -> str:
    values = (
        snapshot_date,
        row.source_row_number,
        row.employer_name,
        row.normalized_employer_name,
        row.trading_name,
        row.normalized_trading_name,
        row.nzbn,
        row.accreditation_type,
        row.accreditation_status,
        row.sector,
        row.subsector,
        row.expiry_date,
        row.start_date,
        row.region,
        row.city,
    )
    return "(" + ", ".join(sql_literal(value) for value in values) + ")"


def generate_sql(
    source_path: Path,
    source_sha256: str,
    snapshot: date,
    rows: list[ImportRow],
    prepared_at: int,
    batch_size: int,
) -> str:
    snapshot_text = snapshot.isoformat()
    snapshot_epoch = int(datetime.combine(snapshot, datetime.min.time(), timezone.utc).timestamp())
    importable_count = sum(row.nzbn is not None for row in rows)
    expected_count = len(rows)
    statements = [
        "-- Generated by scripts/prepare-official-import.py. Do not edit by hand.\n",
        f"-- Source: {source_path.name}\n",
        f"-- SHA-256: {source_sha256}\n",
        f"-- Snapshot: {snapshot_text}; rows: {expected_count}; importable NZBNs: {importable_count}\n\n",
        statement("PRAGMA foreign_keys = ON"),
        statement(
            """
            INSERT INTO official_employer_imports (
              snapshot_date, source_filename, source_sha256,
              expected_row_count, importable_row_count, actual_row_count,
              status, prepared_at, activated_at
            ) VALUES ({snapshot}, {filename}, {sha256}, {expected}, {importable}, 0,
                      'loading', {prepared_at}, NULL)
            ON CONFLICT(snapshot_date) DO UPDATE SET
              source_filename = excluded.source_filename,
              source_sha256 = excluded.source_sha256,
              expected_row_count = excluded.expected_row_count,
              importable_row_count = excluded.importable_row_count,
              actual_row_count = 0,
              status = 'loading',
              prepared_at = excluded.prepared_at,
              activated_at = NULL
            """.format(
                snapshot=sql_literal(snapshot_text),
                filename=sql_literal(source_path.name),
                sha256=sql_literal(source_sha256),
                expected=expected_count,
                importable=importable_count,
                prepared_at=prepared_at,
            )
        ),
        statement(
            f"DELETE FROM official_employer_import_rows WHERE snapshot_date = {sql_literal(snapshot_text)}"
        ),
    ]

    columns = """
      snapshot_date, source_row_number,
      employer_name, normalized_employer_name,
      trading_name, normalized_trading_name, nzbn,
      accreditation_type, accreditation_status,
      sector, subsector,
      expiry_date_of_accreditation, accreditation_start_date,
      region, city
    """.strip()
    for offset in range(0, len(rows), batch_size):
        values = ",\n".join(row_values(snapshot_text, row) for row in rows[offset : offset + batch_size])
        statements.append(
            statement(
                f"INSERT INTO official_employer_import_rows ({columns}) VALUES\n{values}"
            )
        )

    statements.extend(
        [
            statement(
                f"""
                UPDATE official_employer_imports
                   SET actual_row_count = (
                         SELECT COUNT(*)
                           FROM official_employer_import_rows
                          WHERE snapshot_date = {sql_literal(snapshot_text)}
                       ),
                       status = 'validated'
                 WHERE snapshot_date = {sql_literal(snapshot_text)}
                """
            ),
            statement(
                f"""
                INSERT INTO employers (
                  employer_name, normalized_employer_name,
                  trading_name, normalized_trading_name,
                  nzbn, expiry_date_of_accreditation,
                  first_seen_at, last_verified_at, last_verified_source,
                  accreditation_type, accreditation_status,
                  sector, subsector, accreditation_start_date,
                  region, city, official_snapshot_date
                )
                SELECT rows.employer_name, rows.normalized_employer_name,
                       rows.trading_name, rows.normalized_trading_name,
                       rows.nzbn, rows.expiry_date_of_accreditation,
                       {snapshot_epoch}, {snapshot_epoch}, 'inz_official_import',
                       rows.accreditation_type, rows.accreditation_status,
                       rows.sector, rows.subsector, rows.accreditation_start_date,
                       rows.region, rows.city, rows.snapshot_date
                  FROM official_employer_import_rows AS rows
                  JOIN official_employer_imports AS imports
                    ON imports.snapshot_date = rows.snapshot_date
                   AND imports.status = 'validated'
                 WHERE rows.snapshot_date = {sql_literal(snapshot_text)}
                   AND rows.nzbn IS NOT NULL
                ON CONFLICT(nzbn) DO UPDATE SET
                  employer_name = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.employer_name ELSE employers.employer_name END,
                  normalized_employer_name = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.normalized_employer_name ELSE employers.normalized_employer_name END,
                  trading_name = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.trading_name ELSE employers.trading_name END,
                  normalized_trading_name = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.normalized_trading_name ELSE employers.normalized_trading_name END,
                  expiry_date_of_accreditation = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.expiry_date_of_accreditation
                    ELSE employers.expiry_date_of_accreditation END,
                  last_verified_at = MAX(employers.last_verified_at, excluded.last_verified_at),
                  last_verified_source = CASE
                    WHEN excluded.last_verified_at >= employers.last_verified_at
                    THEN excluded.last_verified_source ELSE employers.last_verified_source END,
                  accreditation_type = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.accreditation_type ELSE employers.accreditation_type END,
                  accreditation_status = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.accreditation_status ELSE employers.accreditation_status END,
                  sector = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.sector ELSE employers.sector END,
                  subsector = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.subsector ELSE employers.subsector END,
                  accreditation_start_date = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.accreditation_start_date ELSE employers.accreditation_start_date END,
                  region = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.region ELSE employers.region END,
                  city = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.city ELSE employers.city END,
                  official_snapshot_date = CASE
                    WHEN employers.official_snapshot_date IS NULL
                      OR excluded.official_snapshot_date >= employers.official_snapshot_date
                    THEN excluded.official_snapshot_date ELSE employers.official_snapshot_date END
                """
            ),
            statement(
                f"""
                UPDATE official_employer_imports
                   SET status = 'ready', activated_at = {prepared_at}
                 WHERE snapshot_date = {sql_literal(snapshot_text)}
                   AND status = 'validated'
                """
            ),
        ]
    )
    return "".join(statements)


def main() -> int:
    args = parse_args()
    if args.batch_size < 1 or args.batch_size > 250:
        raise ValueError("--batch-size must be between 1 and 250")
    source_path = args.workbook.expanduser().resolve()
    if source_path.suffix.lower() != ".xlsx" or not source_path.is_file():
        raise ValueError(f"Workbook does not exist or is not .xlsx: {source_path}")

    source_bytes = source_path.read_bytes()
    source_sha256 = hashlib.sha256(source_bytes).hexdigest()
    snapshot, rows, header_row_number = read_workbook(source_path)
    prepared_at = int(time.time())
    sql = generate_sql(
        source_path,
        source_sha256,
        snapshot,
        rows,
        prepared_at,
        args.batch_size,
    )
    output_path = args.output.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(sql, encoding="utf-8", newline="\n")

    missing_nzbn = [
        {"sourceRow": row.source_row_number, "employerName": row.employer_name}
        for row in rows
        if row.nzbn is None
    ]
    print(
        json.dumps(
            {
                "output": str(output_path),
                "sourceSha256": source_sha256,
                "snapshotDate": snapshot.isoformat(),
                "headerRow": header_row_number,
                "sourceRows": len(rows),
                "importableEmployers": len(rows) - len(missing_nzbn),
                "auditOnlyRowsWithoutNzbn": missing_nzbn,
                "sqlBytes": len(sql.encode("utf-8")),
                "batchSize": args.batch_size,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError) as error:
        print(f"error: {error}", file=sys.stderr)
        raise SystemExit(1) from error
